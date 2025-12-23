"""
Convert product_history.jsonl to an Excel table.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Dict, Iterable, List, Any


def _iter_history_rows(history_path: str | Path) -> Iterable[Dict[str, Any]]:
    path = Path(history_path)
    with path.open("r", encoding="utf-8") as handle:
        for line_no, line in enumerate(handle, 1):
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                print(f"[WARN] Skipping invalid JSON at line {line_no}")
                continue

            products = record.get("products") or []
            if not isinstance(products, list):
                continue

            snapshot_ts = record.get("ts") or ""
            region = record.get("region") or ""
            signature = record.get("signature") or ""
            count = record.get("count") if record.get("count") is not None else len(products)

            for product in products:
                if not isinstance(product, dict):
                    continue
                yield {
                    "snapshot_ts": snapshot_ts,
                    "region": region,
                    "product_count": count,
                    "signature": signature,
                    "name": product.get("name") or "",
                    "color": product.get("color") or "",
                    "price": product.get("price") or "",
                    "unavailable": product.get("unavailable"),
                    "url": product.get("url") or "",
                    "is_bag": product.get("is_bag"),
                }


def convert_product_history_to_excel(
    history_path: str | Path = "output/product_history.jsonl",
    output_path: str | Path = "output/product_history.xlsx",
) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise SystemExit("openpyxl is required: pip install openpyxl") from exc

    headers = [
        "snapshot_ts",
        "region",
        "product_count",
        "signature",
        "name",
        "color",
        "price",
        "unavailable",
        "url",
        "is_bag",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "history"
    ws.append(headers)

    row_count = 1
    for row in _iter_history_rows(history_path):
        ws.append([row.get(h) for h in headers])
        row_count += 1

    if row_count > 1:
        last_col = get_column_letter(len(headers))
        table_ref = f"A1:{last_col}{row_count}"
        table = Table(displayName="HistoryTable", ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert product history JSONL to Excel")
    parser.add_argument(
        "history_path",
        nargs="?",
        default="output/product_history.jsonl",
        help="Path to product_history.jsonl",
    )
    parser.add_argument(
        "output_path",
        nargs="?",
        default="output/product_history.xlsx",
        help="Path to output .xlsx",
    )
    args = parser.parse_args()
    output = convert_product_history_to_excel(args.history_path, args.output_path)
    print(f"[INFO] Wrote {output}")


if __name__ == "__main__":
    main()
